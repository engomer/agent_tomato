import RPi.GPIO as GPIO
import serial
import time,sys
import smbus
import math
from datetime import date
from datetime import datetime, timedelta
from time import gmtime,strftime
import time
import openpyxl

excelPATH = "PATH TO EXCEL FILE"

modulePin = 11 #SIM MODULE CONTROL PIN
#GPIO 17

GPIO.setmode(GPIO.BOARD)
GPIO.setup(modulePin, GPIO.OUT)
power_mgmt_1 = 0x6b
power_mgmt_2 = 0x6c

SERIAL_PORT = '/dev/ttyS0'

shockDetectorPin= 0

port = serial.Serial(port='/dev/ttyS0',
   baudrate = 9600,
   parity=serial.PARITY_NONE,
   stopbits=serial.STOPBITS_ONE,
   bytesize=serial.EIGHTBITS,
   timeout=1)


bus = smbus.SMBus(1)
imuaddr = 0x68 #MPU6050 address
shtaddr = 0x44

time.sleep(0.01) #This is so the sensor has tme to preform the mesurement and write its registers before you read it

def setGPIOAnalogRead(shockDetectorPin):
    GPIO.setmode(GPIO.BCM)
    # Set up input pins.
    GPIO.setup(shockDetectorPin, GPIO.IN)

def readShock(shockDetectorPin):

    shockval = GPIO.input(shockDetectorPin)
    print("Shock val is")
    print(shockval)

def readSHT():
    bus.write_i2c_block_data(shtaddr, 0x2C, [0x06])
    time.sleep(0.05)
    data = bus.read_i2c_block_data(0x44, 0x00, 6)

    temp = data[0] * 256 + data[1]
    cTemp = -45 + (175 * temp / 65535.0)
    fTemp = -49 + (315 * temp / 65535.0)
    humidity = 100 * (data[3] * 256 + data[4]) / 65535.0
    cTemp = round(cTemp,2)
    humidity = round(humidity,2)
    # Output data to screen
    print ("Temperature in Celsius is : %.2f C" %cTemp)
    print ("Temperature in Fahrenheit is : %.2f F" %fTemp)
    print ("Relative Humidity is : %.2f %%RH" %humidity)
    
    return [cTemp, humidity]

    
def read_byte(adr):
    return bus.read_byte_data(imuaddr, adr)

def read_word(adr):
    high = bus.read_byte_data(imuaddr, adr)
    low = bus.read_byte_data(imuaddr, adr+1)
    val = (high << 8) + low
    return val

def read_word_2c(adr):
    val = read_word(adr)
    if (val >= 0x8000):
    	return -((65535 - val) + 1)
    else:
    	return val

def dist(a,b):
    return math.sqrt((a*a)+(b*b))

def get_y_rotation(x,y,z):
    radians = math.atan2(x, dist(y,z))
    return math.degrees(radians)

def get_x_rotation(x,y,z):
    radians = math.atan2(y, dist(x,z))
    return -math.degrees(radians)


def getIMU():
    bus.write_byte_data(imuaddr, power_mgmt_1, 0)
    time.sleep(0.05)
    
    gyro_xout = read_word_2c(0x43)
    gyro_yout = read_word_2c(0x45)
    gyro_zout = read_word_2c(0x47)
    
    accel_xout = read_word_2c(0x3b)
    accel_yout = read_word_2c(0x3d)
    accel_zout = read_word_2c(0x3f)

    accel_xout_scaled = accel_xout / 16384.0
    accel_yout_scaled = accel_yout / 16384.0
    accel_zout_scaled = accel_zout / 16384.0
    
    arr = [round(gyro_xout/113 , 3) , round(gyro_yout/113 , 3), round(gyro_zout/113, 3), round(accel_xout_scaled,3) , round(accel_yout_scaled, 3), round(accel_zout_scaled, 3), round(get_x_rotation(accel_xout_scaled, accel_yout_scaled, accel_zout_scaled), 3), round(get_y_rotation(accel_xout_scaled, accel_yout_scaled, accel_zout_scaled), 3)]

    return arr

    time.sleep(0.1)

def powerGSM():
    answer = sendCommand("AT", "OK")
    if answer == 0:
        GPIO.output(modulePin, GPIO.HIGH)
        time.sleep(3)
        GPIO.output(modulePin, GPIO.LOW)
        time.sleep(6)
        answer = sendCommand("AT", "OK")
        if answer == 0:
            powerGSM()



def sendCommand(command,response):
    ans = 0
    c = command + "\r\n"
    port.write(c.encode())
    time.sleep(0.4)
    answer = port.read(500)
    
    a =bytes(command + "\r\r\n" + response + "\r\n", 'utf-8')
    print(answer)
    print(a)
    if a==answer :
        ans = 1
    else:
        ans = 0
    return ans
    


def sendCommand2(command,response, response2):
    ans = 0
    port.write(command + '\r\n')
    answer = port.read(15)

    if answer == response:
        answer = port.read(15)
        if answer == response2:
            ans = 1
    else:
        ans = 0
    return ans

def pushData(data):
    
    setGPRS()
    url = "http://tomatotrack.herokuapp.com/api/add.php?"
    data[0] = data[0].replace("/","%2F")
    ddate = "d=" + str(data[0])
    data[1] = data[1].replace(":","%3A")
    ttime = "&t=" + str(data[1])
    temp = "&temp=" + str(data[2])
    hum = "&hum=" + str(data[3])
    jrx = "&x=" + str(data[4])
    jry = "&y=" + str(data[5])
    jrz = "&z=" + str(data[6])
    accx = "&ax=" + str(data[7])
    accy = "&ay=" + str(data[8])
    accz = "&az=" + str(data[9])
    rotx = "&rx=" + str(data[10])
    roty = "&ry=" + str(data[11])
    shock = "&s=" + str(data[12])
    lat = "&la=" + str(data[13])
    longg = "&lo=" + str(data[14])
    
    resp = sendCommand("AT+HTTPINIT", "OK")
    while resp==0:
	    resp = sendCommand("AT+HTTPINIT", "OK")
	    resp=1
    if resp == 1:
        resp = sendCommand("AT+HTTPPARA=\"CID\",1", "OK")
        if resp == 1:
            resp = sendCommand("AT+HTTPPARA=\"URL\"," + "\"" + url + ddate + ttime + temp + hum + jrx + jry + jrz + accx + accy + accz + rotx + roty +shock + lat + longg + "\"" , "OK")
            if resp == 1:
                resp = sendCommand("AT+HTTPACTION=0" , "OK")
                time.sleep(3)
                if resp == 1:
                    c = "AT+HTTPREAD=1,100000" + '\r\n'
                    time.sleep(6)
                    port.write(c.encode())
                    a = port.read(1500)
                    a = a.decode("utf-8")
                    st = a.count("Row Posted")
                    print(st)
                    if st == 1:
                       return 1
                    else:
                        return 0
                    time.sleep(6)
                else:
                    print("Error getting the url\n")
            else:
                print("Error setting the url\n")
        else:
            print("Error setting the CID")
    else:
        print("Error on initialization")

    sendCommand("AT+HTTPTERM", "OK")

    time.sleep(0.2)

    closeGate()

def setGPRS():

    r = sendCommand("AT+SAPBR=3,1,\"Contype\",\"GPRS\"", "OK")
    while r==0:
        r = sendCommand("AT+SAPBR=3,1,\"Contype\",\"GPRS\"", "OK")
	
    r = sendCommand("AT+SAPBR=3,1,\"APN\",\"INTERNET\"","OK")
    while r==0:
        r = sendCommand("AT+SAPBR=3,1,\"APN\",\"INTERNET\"","OK")
    r = sendCommand("AT+SAPBR=1,1", "OK")
    r = sendCommand("AT+SAPBR=1,1", "OK")
    r = sendCommand("AT+SAPBR=1,1", "OK")
    r = sendCommand("AT+SAPBR=1,1", "OK")
    r = sendCommand("AT+SAPBR=1,1", "OK")
    r = sendCommand("AT+SAPBR=1,1", "OK")
    r = sendCommand("AT+SAPBR=1,1", "OK")
    r = sendCommand("AT+SAPBR=1,1", "OK")

def closeGate():
    sendCommand("AT+SAPBR=0,1", "OK")

def getGPS():
    #get gps long/lat data from sim module
    
    resp = sendCommand("AT+CGNSPWR=1", "OK")
    if resp == 1:
        resp = sendCommand('AT+CGNSSEQ="RMC"',"OK")
        if resp == 1: 
            a = "AT+CGNSINF\r\n"
            port.write(a.encode())
            resp = port.read(150)
            arr = resp.decode("utf-8").split(',')
            
            lat=arr[3]
            longt=arr[4]
            
            return [lat , longt]
            
            
            
        else:
            print("CGNSSEQ error")
    else:
        print("CGNSPWR Error")

def getDateandTime():
    #returns local date and time
    today = date.today()
    d1 =today.strftime("%d/%m/%Y")

    
    #current_time = time.strftime("%H:%M:%S", time.gmtime()) + timedelta(hours=3)
    current_time = datetime.now() #+ timedelta(hours=3)
    c =format(current_time, "%H:%M:%S")
    return [d1 , c]

def parseExcel(ddate,ttime,temperaturee,humidityy,imu_jr_x,imu_jr_y,imu_jr_z,imu_acc_x,imu_acc_y,imu_acc_z,imu_rot_x,imu_rot_y,shock,gps_lattitude,gps_longtitude):
    #puts values to Excel
    file = openpyxl.load_workbook(excelPATH)
    sheet = file['sheet']

    roww = ((ddate,ttime,temperaturee,humidityy,imu_jr_x,imu_jr_y,imu_jr_z,imu_acc_x,imu_acc_y,imu_acc_z,imu_rot_x,imu_rot_y,shock,gps_lattitude,gps_longtitude, "0"))

    #for row in roww:
    sheet.append(roww)

    file.save(excelPATH)

def checkExcel(): #checks the row written or not
    path = excelPATH
    a = []
    # workbook object is created 
    wb_obj = openpyxl.load_workbook(path) 
  
    sheet_obj = wb_obj.active 

    max_col = sheet_obj.max_row 
  
    # Will print a particular row value 
    for i in range(2, max_col + 1): 
        cell_obj = sheet_obj.cell(row = i, column = 16) 
        if cell_obj.value == "0":
            print("yes")
            rown = i

            for j in range(1,16):
                a.append(sheet_obj.cell(row =  i, column = j).value)
            re = pushData(a)
            if re == 1:
                sheet_obj.cell(row = rown ,column = 16).value = '1'
                wb_obj.save(excelPATH)
            break


def correctGPS(): #checks the blanks gps cells and fills it with previous gps data
    # Give the location of the file 
    path = excelPATH
    # workbook object is created 
    wb_obj = openpyxl.load_workbook(path) 

    sheet_obj = wb_obj.active 

    max_row = sheet_obj.max_row 
  
    lat = sheet_obj.cell(row = 14, column=14)
    longg = sheet_obj.cell(row = 14, column=15)



    # Will print a particular row value 
    for i in range(2, max_row + 1): 
        lat = sheet_obj.cell(row = i, column=14)
        longg = sheet_obj.cell(row = i, column=15) 
        if lat.value is None:
            print("New Latitude value is : ")
            print(sheet_obj.cell(row = i - 1, column = 14).value)
            lat.value = sheet_obj.cell(row = i - 1, column = 14).value
        if longg.value is None:
            print("New longtitude value is : ")
            print(sheet_obj.cell(row= i - 1, column = 15).value)
            longg.value = sheet_obj.cell(row= i - 1, column = 15).value
    
    wb_obj.save(excelPATH)                
    
    
def chkGSM():
    resp = sendCommand("AT","OK")

    if resp == 0:
        powerGSM()
        time.sleep(3)

powerGSM()
time.sleep(5)
[aaaa, baaaa] = getGPS()
time.sleep(2)
[aaaa, baaaa] = getGPS()
time.sleep(2)
while 1:
    try:
        
        [T ,H] = readSHT()
        chkGSM()
        print(T)
        print(H)
        [lat, longg] = getGPS()
        time.sleep(0.15)
        chkGSM()
        time.sleep(0.15)
        print(lat)
        print(longg)
        [xout , yout, zout, acc_xout, acc_yout, acc_zout, rotx, roty] = getIMU()
        time.sleep(0.15)
        chkGSM()
        time.sleep(0.15)
        print(xout)
        print(yout)
        print(zout)
        print(acc_xout)
        print(acc_yout)
        print(acc_zout)
        print(rotx)
        print(roty)
        [d,ti] = getDateandTime()
        time.sleep(0.15)
        chkGSM()
        time.sleep(0.15)
        print(d)
        print(ti)
        time.sleep(0.5)
        parseExcel(d,ti,T,H,xout,yout,zout,acc_xout,acc_yout,acc_zout,rotx,roty,"0",lat,longg)
        time.sleep(0.15)
        chkGSM()
        
        time.sleep(5)
        checkExcel()
        time.sleep(0.5)
        chkGSM()
        time.sleep(6)
        closeGate()
        chkGSM()
        time.sleep(3)
        correctGPS()
        time.sleep(0.15)
        chkGSM()
        time.sleep(5)
        
        
        
    except:
        print("error")
